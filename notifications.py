# Copyright (c) Streamlit Inc. (2018-2022) Snowflake Inc. (2022)
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

import streamlit as st
import pandas as pd
from openpyxl import load_workbook
 
def main():
    st.title("Excel Table Filtering App")
    
    # File upload
    uploaded_file = st.sidebar.file_uploader("Choose an Excel file", type=["xlsx", "xls"])
    
    if uploaded_file is not None:
        # Load the workbook
        wb = load_workbook(uploaded_file)
        sheet_name = wb.sheetnames[0]  # Assume the table is in the first sheet
        ws = wb[sheet_name]
        
        # Find the table range
        table_range = None
        for cell in ws["A"]:
            if cell.value == "Table1":  # Assuming the table has a name "Table1"
                table_range = ws[cell.row+1:ws.max_row]
                break
        
        if table_range:
            # Extract table data into a DataFrame
            data = []
            for row in table_range:
                data.append([cell.value for cell in row])
            df = pd.DataFrame(data[1:], columns=data[0])
            
            # Get unique values for each column
            unique_values = {}
            for column in df.columns:
                unique_values[column] = df[column].unique()
            
            # Multiselect filters
            selected_filters = {}
            for column, values in unique_values.items():
                selected_values = st.sidebar.multiselect(f"Select {column}", values)
                if selected_values:
                    selected_filters[column] = selected_values
            
            # Filter DataFrame based on selected filters
            if selected_filters:
                filter_conditions = []
                for column, values in selected_filters.items():
                    filter_conditions.append(df[column].isin(values))
                
                filtered_df = df[pd.concat(filter_conditions, axis=1).all(axis=1)]
            else:
                filtered_df = df
            
            # Display filtered results in a table
            if not filtered_df.empty:
                st.subheader("Filtered Results")
                st.table(filtered_df)
                
                # Display totals
                totals = filtered_df.sum(numeric_only=True)
                st.subheader("Totals")
                st.table(totals)
            else:
                st.warning("No data matches the selected filters.")
        else:
            st.warning("No table found in the Excel file.")
    
    else:
        st.info("Please upload an Excel file.")
 
if __name__ == "__main__":
    main()
